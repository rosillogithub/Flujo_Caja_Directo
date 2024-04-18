from string import ascii_uppercase
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Alignment, Side, PatternFill
from arrow import utcnow   # utcnow: función que devuelve un objeto `Arrow` que representa la fecha y hora actuales en formato UTC.
from functools import reduce  

# Función para aplicar formato a una celda
def aplicar_formato_celda(celda, valor, estilo):
    celda.value = valor
    celda.font = estilo["font"]
    celda.border = estilo["border"]
    celda.alignment = estilo["alignment"]
    celda.fill = estilo["fill"]

# Función para generar un libro de trabajo y una hoja
def generar_libro_hoja(titulo):                 # Se recibe el título de la hoja
    libro = Workbook()                          # Se crea un libro de trabajo
    hoja = libro.active                         # Se obtiene la hoja activa
    hoja.title = titulo                         # Se asigna el título a la hoja
    hoja.sheet_properties.tabColor = "1072BA"   # Se asigna un color a la pestaña
    hoja.sheet_view.showGridLines = False       # Se ocultan las líneas de la cuadrícula
    return libro, hoja                          # Se retorna el libro y la hoja

# Función para aplicar bordes a una fila
def aplicar_bordes_fila(fila, estilo):      # Se recibe la fila y el estilo
    for celda in fila:                      # Se recorre cada celda de la fila
        celda.border = estilo["border"]     # Se asigna el estilo de borde a la celda

# Función para ajustar el ancho de las columnas
def ajustar_ancho_columnas(hoja):                                                    # Se recibe la hoja
    for columna in hoja.columns:                                                     # Se recorre cada columna de la hoja
        longitud_maxima = reduce(lambda x, y: max(x, len(str(y.value))), columna, 0) # Se obtiene la longitud máxima de los valores de la columna
        ajuste_ancho = (longitud_maxima + 1) * 1.2                                   # Se calcula el ancho de la columna
        hoja.column_dimensions[columna[0].letter].width = ajuste_ancho               # Se ajusta el ancho de la columna

# Función para generar un reporte en Excel
def generar_reporte(titulo, cabecera, registros, nombre_excel): # Se reciben el título, la cabecera, los registros y el nombre del archivo
    thin = Side(border_style="thin", color="000000")            # Se define un borde delgado
    
    # Estilos para el título
    estilo_titulo = {                                                   # Estilo para el título                        
        "font": Font(color="FF000000", size=11, bold=True),             # Fuente
        "border": None,                                                 # Borde
        "alignment": Alignment(horizontal="center", vertical="center"), # Alineación
        "fill": None                                                    # Relleno
    }

    estilo_cabecera = {                                                 # Estilo para la cabecera
        "font": Font(color="FF000000", size=10, bold=True),
        "border": Border(top=thin, left=thin, right=thin, bottom=thin),
        "alignment": Alignment(horizontal="center", vertical="center"),
        "fill": PatternFill("solid", fgColor="C0C0C0")
    }

    estilo_registro = {                                                     # Estilo para los registros
        "font": Font(color="FF000000", size=10, bold=False),
        "border": Border(top=thin, left=thin, right=thin, bottom=thin),
        "alignment": Alignment(horizontal="left", vertical="center"),
        "fill": None
    }

    libro, hoja = generar_libro_hoja(titulo) # Se obtiene el libro y la hoja

    # Se obtiene la celda de la hoja
    celda_titulo = hoja.cell(row=2, column=2)                               # Se obtiene la celda de la hoja
    aplicar_formato_celda(celda_titulo, titulo.upper(), estilo_titulo)      # Se aplica el formato a la celda

    # Información extra
    info_extra = [                                                          # Información adicional
        ("Generado por: Andres Niño", 5),
        ("Fecha de descarga: {}".format(utcnow().to("local").format("DD/MM/YYYY")), 6),
        ("Registros descargados: {}".format(len(registros)), 8)
    ]
    for info in info_extra:                                                 # Se recorre la información adicional 
        celda_info = hoja.cell(row=info[1], column=2)                       # Se obtiene la celda de la hoja
        aplicar_formato_celda(celda_info, info[0], estilo_registro)         # Se aplica el formato a la celda

    # Cabecera
    for idx, dato in enumerate(cabecera, start=2):                          # Se recorre la cabecera
        celda_cabecera = hoja.cell(row=10, column=idx)                      # Se obtiene la celda de la hoja
        aplicar_formato_celda(celda_cabecera, dato, estilo_cabecera)        # Se aplica el formato a la celda

    # Registros
    for fila_idx, registro in enumerate(registros, start=11):               # Se recorren los registros
        for col_idx, dato in enumerate(registro, start=2):                  # Se recorre cada dato del registro
            celda_registro = hoja.cell(row=fila_idx, column=col_idx)        # Se obtiene la celda de la hoja
            aplicar_formato_celda(celda_registro, dato, estilo_registro)    # Se aplica el formato a la celda

    # Aplicar bordes y ajustar ancho de columnas
    for rango in [hoja["B2:{}3".format(ascii_uppercase[len(cabecera)])], hoja["B10:{}10".format(ascii_uppercase[len(cabecera)])]]: # Recorre rango de celdas
        for fila in rango:                                      # Recorre cada fila del rango
            aplicar_bordes_fila(fila, estilo_cabecera)          # Aplica bordes a la fila
    ajustar_ancho_columnas(hoja)                                # Ajusta el ancho de las columnas

    try:
        libro.save("{}.xlsx".format(nombre_excel))
        return "Reporte generado con éxito."
    except PermissionError:
        return "Error inesperado: Permiso denegado."
    except Exception as e:
        return f"Error desconocido: {str(e)}"
    finally:
        libro.close()

# ==============================================================================================================

# Datos del reporte
titulo = "LISTADO DE USUARIOS"
cabecera = ("D.N.I", "NOMBRE", "APELLIDO", "FECHA DE NACIMIENTO")
registros = [
    (1110800310, "Andres", "Niño", "06/06/2019"),
    (1110800311, "Andres", "Niño", "06/06/2019"),
    (1110800312, "Andres", "Niño", "06/06/2019")
]
nombre_excel = "Listado de usuarios"

# Generar reporte
print(generar_reporte(titulo, cabecera, registros, nombre_excel))

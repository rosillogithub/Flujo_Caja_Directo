from string import ascii_uppercase
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Alignment, Side, PatternFill
from arrow import utcnow   # utcnow: función que devuelve un objeto `Arrow` que representa la fecha y hora actuales en formato UTC.
from functools import reduce  

# Función para aplicar formato a una celda
def aplicar_formato_celda(celda, valor, estilo):
    celda.value = valor
    celda.font = estilo["font"]
    celda.border = estilo["border"]
    celda.alignment = estilo["alignment"]
    celda.fill = estilo["fill"]

# Función para generar un libro de trabajo y una hoja
def generar_libro_hoja(titulo):
    libro = Workbook()
    hoja = libro.active
    hoja.title = titulo
    hoja.sheet_properties.tabColor = "1072BA"
    hoja.sheet_view.showGridLines = False
    return libro, hoja

# Función para aplicar bordes a una fila
def aplicar_bordes_fila(fila, estilo):
    for celda in fila:
        celda.border = estilo["border"]

# Función para ajustar el ancho de las columnas
def ajustar_ancho_columnas(hoja):
    for columna in hoja.columns:
        longitud_maxima = reduce(lambda x, y: max(x, len(str(y.value))), columna, 0)
        ajuste_ancho = (longitud_maxima + 1) * 1.2
        hoja.column_dimensions[columna[0].letter].width = ajuste_ancho


# Función para generar un reporte en Excel
def generar_reporte(titulo, cabecera, registros, nombre_excel):
    thin = Side(border_style="thin", color="000000")
    
    estilo_titulo = {
        "font": Font(color="FF000000", size=11, bold=True),
        "border": None,
        "alignment": Alignment(horizontal="center", vertical="center"),
        "fill": None
    }

    estilo_cabecera = {
        "font": Font(color="FF000000", size=10, bold=True),
        "border": Border(top=thin, left=thin, right=thin, bottom=thin),
        "alignment": Alignment(horizontal="center", vertical="center"),
        "fill": PatternFill("solid", fgColor="C0C0C0")
    }

    estilo_registro = {
        "font": Font(color="FF000000", size=10, bold=False),
        "border": Border(top=thin, left=thin, right=thin, bottom=thin),
        "alignment": Alignment(horizontal="left", vertical="center"),
        "fill": None
    }

    libro, hoja = generar_libro_hoja(titulo)

    # Título
    celda_titulo = hoja.cell(row=2, column=2)
    aplicar_formato_celda(celda_titulo, titulo.upper(), estilo_titulo)

    # Información extra
    info_extra = [
        ("Generado por: Andres Niño", 5),
        ("Fecha de descarga: {}".format(utcnow().to("local").format("DD/MM/YYYY")), 6),
        ("Registros descargados: {}".format(len(registros)), 8)
    ]
    for info in info_extra:
        celda_info = hoja.cell(row=info[1], column=2)
        aplicar_formato_celda(celda_info, info[0], estilo_registro)

    # Cabecera
    for idx, dato in enumerate(cabecera, start=2):
        celda_cabecera = hoja.cell(row=10, column=idx)
        aplicar_formato_celda(celda_cabecera, dato, estilo_cabecera)

    # Registros
    for fila_idx, registro in enumerate(registros, start=11):
        for col_idx, dato in enumerate(registro, start=2):
            celda_registro = hoja.cell(row=fila_idx, column=col_idx)
            aplicar_formato_celda(celda_registro, dato, estilo_registro)

    # Aplicar bordes y ajustar ancho de columnas
    for rango in [hoja["B2:{}3".format(ascii_uppercase[len(cabecera)])], hoja["B10:{}10".format(ascii_uppercase[len(cabecera)])]]:
        for fila in rango:
            aplicar_bordes_fila(fila, estilo_cabecera)
    ajustar_ancho_columnas(hoja)

    try:
        libro.save("{}.xlsx".format(nombre_excel))
        return "Reporte generado con éxito."
    except PermissionError:
        return "Error inesperado: Permiso denegado."
    except Exception as e:
        return f"Error desconocido: {str(e)}"
    finally:
        libro.close()

# Datos del reporte
titulo = "LISTADO DE USUARIOS"
cabecera = ("D.N.I", "NOMBRE", "APELLIDO", "FECHA DE NACIMIENTO")
registros = [
    (1110800310, "Andres", "Niño", "06/06/2019"),
    (1110800311, "Andres", "Niño", "06/06/2019"),
    (1110800312, "Andres", "Niño", "06/06/2019")
]
nombre_excel = "Listado de usuarios"

# Generar reporte
print(generar_reporte(titulo, cabecera, registros, nombre_excel))
